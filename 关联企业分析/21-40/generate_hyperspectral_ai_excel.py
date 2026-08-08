from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parent
SOURCE_FILE = BASE_DIR / "企业高光谱AI业务分析报告-核验版.md"
OUTPUT_FILE = BASE_DIR / "关联企业事业线-高光谱AI能力汇报表.xlsx"


def parse_report() -> list[dict[str, str]]:
    text = SOURCE_FILE.read_text(encoding="utf-8")
    section = text.split("## 四、逐企业、逐业务线分析", 1)[1].split("## 五、建议的首批六个项目", 1)[0]
    company_pattern = re.compile(
        r"^## (?P<number>\d+)\. (?P<company>.+?)（(?P<code>.+?)）\s*$",
        re.MULTILINE,
    )
    matches = list(company_pattern.finditer(section))
    rows: list[dict[str, str]] = []

    for index, match in enumerate(matches):
        block_end = matches[index + 1].start() if index + 1 < len(matches) else len(section)
        block = section[match.end():block_end]
        basic_match = re.search(r"基础信息：(.+)", block)
        basic_info = basic_match.group(1).strip() if basic_match else ""

        for line in block.splitlines():
            line = line.strip()
            if not line.startswith("- ") or line.startswith("- “"):
                continue
            item_match = re.match(r"- (?P<line>.+?)（(?P<fit>.+?)）：(?P<analysis>.+)", line)
            if not item_match:
                continue
            rows.append(
                {
                    "序号": match.group("number"),
                    "公司": match.group("company"),
                    "证券代码": match.group("code"),
                    "企业基础信息": basic_info,
                    "事业线": item_match.group("line").strip(),
                    "原始适配判断": item_match.group("fit").strip(),
                    "原报告分析": item_match.group("analysis").strip(),
                }
            )
    return rows


def opportunity_grade(fit: str) -> str:
    if "低至中" in fit or ("高" in fit and "低" in fit) or ("中" in fit and "低" in fit):
        return "B-验证型"
    if "高" in fit:
        return "A-优先"
    if "中" in fit:
        return "B-验证型"
    return "C-储备"


def infer_capabilities(line_name: str, analysis: str, grade: str) -> tuple[str, str, str, str, str]:
    context = f"{line_name}{analysis}"

    if any(word in context for word in ["玻璃", "透射", "反射", "镀膜", "膜材", "膜层", "光伏"]):
        hyperspectral = "透射/反射高光谱成像；膜层与光谱性能标定；线扫采集、标准光源及在线校准"
    elif any(word in context for word in ["晶圆", "外延", "光刻胶", "半导体", "Micro", "Mini-LED", "显微"]):
        hyperspectral = "显微高光谱成像；微区光谱采集；膜厚、残胶、污染及发光一致性检测"
    elif any(word in context for word in ["甲烷", "气体", "遥测", "矿产", "机载遥感"]):
        hyperspectral = "VNIR/SWIR高光谱遥感；特征吸收峰识别；大范围扫描与辐射/几何校准"
    elif any(word in context for word in ["分选", "回收", "塑料", "材料", "粉体", "浆料", "矿物", "异物"]):
        hyperspectral = "VNIR/SWIR线扫高光谱；材料光谱指纹库；输送带在线采集与分选联动"
    elif any(word in context for word in ["胶", "油", "污染", "清洁", "涂层", "水分", "残留", "绝缘"]):
        hyperspectral = "VNIR/SWIR高光谱成像；有机物、水分、污染及涂层差异检测；在线标定"
    elif any(word in context for word in ["电源", "信号链", "ADC", "AFE", "MCU", "存储", "网络", "控制板"]):
        hyperspectral = "高光谱相机与整机平台；探测器采集、标定、时序同步及边缘数据接口能力"
    else:
        hyperspectral = "VNIR/SWIR高光谱成像；光谱采集标定；线扫/面阵设备集成与边缘数据输出"

    if any(word in context for word in ["机器人", "分选", "剔除", "机械臂"]):
        ai = "光谱分类/目标定位；RGB+光谱多模态融合；机器人/气吹策略控制；边缘实时推理"
        product = "高光谱AI分选工作站"
        poc = "建立3—5类物料光谱库，完成识别、定位、分拣闭环，验证准确率、漏检率和节拍"
    elif any(word in context for word in ["电源", "信号链", "ADC", "AFE", "MCU", "存储", "网络", "控制板", "光源"]):
        ai = "采集控制SDK；暗电流/温漂/坏点校正；数据压缩与边缘预处理；设备健康监测"
        product = "高光谱核心部件参考设计与SDK"
        poc = "完成样机适配、噪声与稳定性测试，输出参考BOM、驱动SDK和联合方案白皮书"
    elif any(word in context for word in ["云", "平台", "数字孪生", "数据集", "授权"]):
        ai = "光谱数据治理；模型训练与版本管理；多源数据融合；API服务、可视化及持续学习"
        product = "光谱AI模型平台/行业数据服务"
        poc = "打通采集、标注、训练、部署和回流链路，验证模型增益、接口性能及运营成本"
    elif any(word in context for word in ["教育", "实训"]):
        ai = "教学样本管理；光谱分类实验；低代码建模；课程任务、报告生成与模型评测"
        product = "高光谱AI教学实训平台"
        poc = "搭建食品、植物、矿物等示范实验，形成课程包、数据集和可复现实验流程"
    elif grade == "C-储备":
        ai = "场景数据评估；小样本建模；可行性验证；必要时与RGB/热像等多模态融合"
        product = "联合预研/场景验证包"
        poc = "先完成需求与样本评估，仅在高光谱相对现有方案有明确增益时进入PoC"
    else:
        ai = "光谱预处理与特征提取；分类/回归/异常检测；小样本学习；边缘部署；MES/QMS集成"
        product = "高光谱AI在线质检系统"
        poc = "采集良品/缺陷真值样本，完成离线建模与盲测，再验证漏检率、误报率、GR&R和产线节拍"

    ai += "；多模态大模型；行业知识库/RAG；质检分析智能体；自然语言报告生成"

    if grade == "A-优先":
        value = "形成可量化质检或分选闭环，适合优先立项并沉淀行业模型"
    elif grade == "B-验证型":
        value = "技术可行但受成本、节拍、曲面或样本规模影响，建议限定工位验证"
    else:
        value = "高光谱不是刚需或仅属通用配套，暂不投入专项产品化资源"

    return hyperspectral, ai, product, poc, value


def enrich_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    result = []
    for row in rows:
        grade = opportunity_grade(row["原始适配判断"])
        hyperspectral, ai, product, poc, value = infer_capabilities(
            row["事业线"], row["原报告分析"], grade
        )
        result.append(
            {
                **row,
                "机会等级": grade,
                "中达瑞和高光谱基础能力": hyperspectral,
                "AI团队应用开发能力": ai,
                "联合方案/应用场景": row["原报告分析"],
                "建议产品形态": product,
                "首期PoC建议": poc,
                "汇报价值判断": value,
            }
        )
    return result


def set_sheet_title(ws, title: str, subtitle: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    ws.cell(1, 1, title)
    ws.cell(1, 1).font = Font(size=18, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor="1F4E78")
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).font = Font(size=10, color="666666")
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 28


def style_data_sheet(ws, widths: list[int], header_row: int, data_start: int) -> None:
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.row_dimensions[header_row].height = 34

    for row in ws.iter_rows(min_row=data_start):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        grade = row[6].value if len(row) > 6 else None
        if grade == "A-优先":
            row[6].fill = PatternFill("solid", fgColor="C6E0B4")
        elif grade == "B-验证型":
            row[6].fill = PatternFill("solid", fgColor="FFE699")
        elif grade == "C-储备":
            row[6].fill = PatternFill("solid", fgColor="E7E6E6")

    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(widths))}{ws.max_row}"


def add_overview(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "汇报总览"
    set_sheet_title(
        ws,
        "关联企业事业线 × 高光谱基础能力 × AI应用开发能力",
        "汇报口径：企业事实来自核验版报告；联合方案均为建议，不代表企业已部署。资料截止：2026-07-20。",
        8,
    )

    companies = sorted({row["公司"] for row in rows}, key=lambda name: int(next(r["序号"] for r in rows if r["公司"] == name)))
    grades = Counter(row["机会等级"] for row in rows)
    metrics = [
        ("覆盖企业", len(companies), "家"),
        ("覆盖事业线", len(rows), "条"),
        ("A-优先机会", grades["A-优先"], "条"),
        ("B-验证型机会", grades["B-验证型"], "条"),
        ("C-储备机会", grades["C-储备"], "条"),
    ]
    for col, (label, value, unit) in enumerate(metrics, start=1):
        ws.cell(4, col, label)
        ws.cell(5, col, value)
        ws.cell(5, col).number_format = f'0"{unit}"'
        ws.cell(4, col).font = Font(bold=True, color="666666")
        ws.cell(5, col).font = Font(size=16, bold=True, color="1F4E78")
        ws.cell(4, col).alignment = ws.cell(5, col).alignment = Alignment(horizontal="center")

    ws["A8"] = "能力组合"
    ws["A8"].font = Font(size=13, bold=True, color="1F4E78")
    capability_rows = [
        ("高光谱硬件底座", "VNIR/SWIR成像、显微/线扫/面阵采集、光源、标定、边缘数据输出"),
        ("AI应用开发", "光谱预处理、分类/回归/异常检测、小样本学习、多模态融合、边缘部署"),
        ("AI大模型能力", "多模态大模型、行业知识库/RAG、质检分析智能体、自然语言交互与报告生成"),
        ("工业化集成", "PLC/机器人/MES/QMS接口、模型版本管理、数据回流与持续学习"),
        ("交付与商业化", "PoC验证、行业光谱库、软硬一体工作站、模型授权和年度运维"),
    ]
    for row_index, values in enumerate(capability_rows, start=9):
        ws.cell(row_index, 1, values[0]).font = Font(bold=True)
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=8)
        ws.cell(row_index, 2, values[1])

    ws["A16"] = "首批建议项目"
    ws["A16"].font = Font(size=13, bold=True, color="1F4E78")
    projects = [
        ("宁德时代 / 比亚迪", "电池极片残余水分与涂层均匀性离线标定PoC"),
        ("福耀玻璃", "高附加值汽车玻璃镀膜与透射光谱在线检测"),
        ("汇川技术", "标准化高光谱AI分选机器人工作站"),
        ("鼎泰高科", "功能膜卷对卷光谱检测及装备选配模块"),
        ("芯碁微装", "光刻胶残留或Mini/Micro-LED光谱一致性检测"),
        ("敏实集团", "电池盒涂胶前洁净度与塑件材料防错"),
    ]
    for row_index, values in enumerate(projects, start=17):
        ws.cell(row_index, 1, values[0]).font = Font(bold=True)
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=8)
        ws.cell(row_index, 2, values[1])

    ws["A25"] = "统一PoC验收指标"
    ws["A25"].font = Font(size=13, bold=True, color="1F4E78")
    ws.merge_cells("A26:H27")
    ws["A26"] = "实验室真值一致性、漏检率、误报率、节拍、GR&R、环境鲁棒性、单线投资回收期、模型年度维护成本；大模型回答准确率、幻觉率、知识引用可追溯性和人工复核效率。"
    ws["A26"].alignment = Alignment(wrap_text=True, vertical="center")
    for col, width in enumerate([22, 25, 25, 25, 25, 16, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=8, max_row=27):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def add_detail(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("事业线机会清单")
    headers = [
        "序号", "公司", "证券代码", "事业线", "企业基础信息", "原始适配判断", "机会等级",
        "中达瑞和高光谱基础能力", "AI团队应用开发能力", "联合方案/应用场景",
        "建议产品形态", "首期PoC建议", "汇报价值判断",
    ]
    set_sheet_title(
        ws,
        "逐公司、逐事业线高光谱AI机会清单",
        "可按公司、机会等级和事业线筛选。A=优先立项；B=限定工位验证；C=储备或不建议专项投入。",
        len(headers),
    )
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col, header)
    for row_index, row in enumerate(rows, start=5):
        for col, header in enumerate(headers, start=1):
            ws.cell(row_index, col, row[header])

    widths = [7, 16, 18, 25, 38, 18, 13, 42, 42, 55, 27, 52, 38]
    style_data_sheet(ws, widths, 4, 5)
    table = Table(displayName="BusinessLineOpportunities", ref=f"A4:M{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def add_company_summary(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("公司汇总")
    headers = ["序号", "公司", "证券代码", "事业线数", "A-优先", "B-验证型", "C-储备", "首选切入事业线", "建议汇报动作"]
    set_sheet_title(
        ws,
        "公司级机会汇总",
        "用于管理层快速确定拜访顺序、首个PoC和资源投入级别。",
        len(headers),
    )
    for col, header in enumerate(headers, start=1):
        ws.cell(4, col, header)

    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[row["公司"]].append(row)

    companies = sorted(grouped, key=lambda name: int(grouped[name][0]["序号"]))
    for row_index, company in enumerate(companies, start=5):
        items = grouped[company]
        counts = Counter(item["机会等级"] for item in items)
        preferred = next((item for item in items if item["机会等级"] == "A-优先"), items[0])
        if counts["A-优先"] >= 2:
            action = "优先拜访，选择一个工位启动付费PoC"
        elif counts["A-优先"] == 1:
            action = "围绕首选事业线开展样本与节拍评估"
        elif counts["B-验证型"] > 0:
            action = "限定场景预研，达到增益门槛后再立项"
        else:
            action = "保持业务跟踪，不投入专项研发资源"
        values = [
            items[0]["序号"], company, items[0]["证券代码"], len(items),
            counts["A-优先"], counts["B-验证型"], counts["C-储备"],
            preferred["事业线"], action,
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_index, col, value)

    widths = [7, 18, 20, 12, 12, 12, 12, 30, 45]
    style_data_sheet(ws, widths, 4, 5)
    for row in ws.iter_rows(min_row=5):
        if row[4].value >= 2:
            row[1].fill = PatternFill("solid", fgColor="C6E0B4")
    table = Table(displayName="CompanySummary", ref=f"A4:I{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def add_reporting_notes(wb: Workbook) -> None:
    ws = wb.create_sheet("汇报口径")
    set_sheet_title(ws, "汇报口径与边界", "防止将技术可行性表述为已落地项目或高光谱刚需。", 6)
    sections = [
        ("建议话术", "我们提供“高光谱硬件底座+AI应用开发+工业系统集成”的联合能力，先以可量化PoC验证业务价值，再推进产品化。"),
        ("事实边界", "企业业务信息来自核验版报告；表中联合方案为合作建议，不代表企业已有采购计划、已部署或已形成收入。"),
        ("技术边界", "高光谱适合识别材料、含水率、膜层、污染和光谱性能；尺寸、内部裂纹、精确元素及电芯内部SOH应结合RGB、3D、X射线、超声、XRF/LIBS或电化学手段。"),
        ("立项原则", "优先选择现有方法难以区分、可获取真值、节拍可验证、ROI可计算的单一工位；避免先建大平台再找场景。"),
        ("商业模式", "PoC服务费 → 软硬一体设备/工作站 → 行业模型授权 → 新物料建模费 → 年度标定与运维服务。"),
        ("数据闭环", "样本采集、真值标注、模型训练、边缘部署、异常复核、数据回流、模型版本管理和效果持续监控。"),
    ]
    for row_index, (title, content) in enumerate(sections, start=4):
        ws.cell(row_index, 1, title).font = Font(bold=True, color="1F4E78")
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=6)
        ws.cell(row_index, 2, content)
        ws.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row_index].height = 55
    ws.column_dimensions["A"].width = 18
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 25


def build_workbook() -> None:
    rows = enrich_rows(parse_report())
    workbook = Workbook()
    add_overview(workbook, rows)
    add_detail(workbook, rows)
    add_company_summary(workbook, rows)
    add_reporting_notes(workbook)
    workbook.save(OUTPUT_FILE)

    # 重新加载，确保生成文件结构完整且可被Excel解析。
    checked = load_workbook(OUTPUT_FILE, read_only=True, data_only=False)
    assert checked.sheetnames == ["汇报总览", "事业线机会清单", "公司汇总", "汇报口径"]
    assert checked["事业线机会清单"].max_row == len(rows) + 4
    assert checked["公司汇总"].max_row == 24
    checked.close()
    print(f"已生成：{OUTPUT_FILE}")
    print(f"企业数：{len({row['公司'] for row in rows})}；事业线数：{len(rows)}")
    print(f"机会等级：{dict(Counter(row['机会等级'] for row in rows))}")


if __name__ == "__main__":
    build_workbook()
